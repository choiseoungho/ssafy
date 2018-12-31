import urllib.request

from bs4 import BeautifulSoup
from openpyxl import Workbook
from elice_utils import EliceUtils

elice_utils = EliceUtils()


def main():
    wb = Workbook()
    
    sheet1 = wb.active
    file_name = 'sample.xlsx'
    
    url = "http://www.newsis.com/eco/list/?cid=10400&scid=10404"
    req = urllib.request.Request(url)
    sourcecode = urllib.request.urlopen(url).read()
    soup = BeautifulSoup(sourcecode, "html.parser")
    
    articles = []
    
    for i in soup.find_all("strong", class_="title"):
        articles.append(i.get_text())
    
    for row_index in range(1,len(articles)+1):
        sheet1.cell(row=row_index, column=1).value = row_index
        sheet1.cell(row=row_index, column=2).value = articles[(row_index-1)]
    
    wb.save(filename=file_name)
    elice_utils.send_file(file_name)


if __name__ == "__main__":
    main()
